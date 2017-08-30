from __future__ import print_function

import datetime
import openpyxl
import os
from time import sleep

try:
    from tango import DeviceProxy, DevState, ApiUtil, DevFailed
except ImportError:
    from PyTango import DeviceProxy, DevState, ApiUtil, DevFailed
from numpy import linspace

from ds_tanks.tanks_utils import plot_with_optimal_trajectories, plot_results


def run_optimisation_through_ds(with_commands=True, with_plots=True,
                                h1f=16, h2f=16, h3f=15, r=0.01, opt_dev=None):
    if opt_dev is None:
        opt_dev = DeviceProxy("opt/ctrl/1")
    if with_commands:
        q = [[0.1, 0.0, 0.0],
             [0.0, 0.1, 0.0],
             [0.0, 0.0, 0.1]]
        print("Setting initial values...")
        print("H1 final:", opt_dev.write_read_attribute("H1Final", h1f).value)
        print("H2 final:", opt_dev.write_read_attribute("H2Final", h2f).value)
        print("H3 final:", opt_dev.write_read_attribute("H3Final", h3f).value)
        print("Control weight:", opt_dev.write_read_attribute("R", r).value)
        print("State weights:", opt_dev.write_read_attribute("Q", q).value)

        print("Optimising...")
        for _ in range(0, 4):
            if opt_dev.state() == DevState.RUNNING:
                sleep(5)
            else:
                break
        opt_dev.command_inout_asynch("Optimise", False)
        sleep(20)

        if opt_dev.state() == DevState.ALARM:
            print("Optimisation failed, exiting!")
            return
        try:
            opt_dev.command_inout("SendControl")
        except DevFailed as e:
            if "Timeout" in e[1].desc:
                print("Waiting for LQR results...")
                sleep(5)
            else:
                print("Something went wrong!")
                print(e)
        # Running verification...
        opt_dev.command_inout("RunVerification")
        sleep(8)

    if opt_dev.state() == DevState.STANDBY:
        opt_control = opt_dev.read_attribute("OptimalControl").value
        opt_time = opt_dev.read_attribute("OptimalTime").value
        time_opt = linspace(0.0, opt_time, len(opt_control))
        opt_h1 = opt_dev.read_attribute("OptimalH1").value
        opt_h2 = opt_dev.read_attribute("OptimalH2").value
        opt_h3 = opt_dev.read_attribute("OptimalH3").value
        if with_plots:
            plot_results(opt_h1, opt_h2, opt_h3, time_opt, opt_control,
                         "Optimised with TanksOptimalControl")

        h1_sim = opt_dev.read_attribute("H1Simulated").value
        h2_sim = opt_dev.read_attribute("H2Simulated").value
        h3_sim = opt_dev.read_attribute("H3Simulated").value
        time_sim = opt_dev.read_attribute("SimulationTime").value
        print(get_matlab_data(opt_dev))
        if with_plots:
            plot_with_optimal_trajectories(time_opt, time_sim, h1_sim, h2_sim,
                                           h3_sim, opt_h1, opt_h2, opt_h3,
                                           opt_control)
    else:
        print("Wrong state, exiting!")


def get_matlab_data(device):
    switch_times = device.read_attribute('SwitchTimes').value
    k = device.read_attribute("K").value
    data = [device.read_attribute('H1Final').value,
            device.read_attribute('H2Final').value,
            device.read_attribute('H3Final').value,
            device.read_attribute('OptimalTime').value,
            device.read_attribute('OptimalControl').value[0],
            0,
            switch_times[0],
            switch_times[1],
            65.9,
            k[0],
            k[1],
            k[2]]
    return data


def log_plotting_data(opt_dev, log_file="data_log.xlsx", title_prefix=""):
    data = get_data_for_log(opt_dev)

    if os.path.isfile(os.path.abspath(log_file)):
        workbook = openpyxl.load_workbook(log_file)
    else:
        workbook = openpyxl.Workbook()
    title = str(datetime.datetime.time(datetime.datetime.now()))
    title = title[:title.rfind('.')].replace(":", "-")
    title = title_prefix + " " + title
    worksheet = workbook.create_sheet(title=title)
    for i in range(len(data)):
        worksheet.append(data[i])
    workbook.save(log_file)
    print("Saved data to '%s' worksheet of '%s'" % (title, log_file))


def get_data_for_log(opt_dev):
    data = []
    if "not found" in opt_dev.status():
        return data
    opt_control = opt_dev.read_attribute("OptimalControl").value
    opt_time = opt_dev.read_attribute("OptimalTime").value
    time_opt = linspace(0.0, opt_time, len(opt_control))
    opt_h1 = opt_dev.read_attribute("OptimalH1").value
    opt_h2 = opt_dev.read_attribute("OptimalH2").value
    opt_h3 = opt_dev.read_attribute("OptimalH3").value
    h1_sim = opt_dev.read_attribute("H1Simulated").value
    h2_sim = opt_dev.read_attribute("H2Simulated").value
    h3_sim = opt_dev.read_attribute("H3Simulated").value
    time_sim = opt_dev.read_attribute("SimulationTime").value
    data.append(opt_control.tolist())
    data.append((opt_time,))
    data.append(time_opt.tolist())
    data.append(opt_h1.tolist())
    data.append(opt_h2.tolist())
    data.append(opt_h3.tolist())
    data.append(h1_sim.tolist())
    data.append(h2_sim.tolist())
    data.append(h3_sim.tolist())
    data.append(time_sim.tolist())
    data.append((opt_dev.read_attribute("VerificationError").value,))
    return data


def run_looped_optimisation(step=1, h_min=1, h_max=40):
    opt_dev = DeviceProxy("opt/ctrl/1")
    for h1_final in range(h_min, h_max, step):
        for h2_final in range(h_min, h_max, step):
            for h3_final in range(h_min, h_max, step):
                run_optimisation_through_ds(True, False, h1_final, h2_final,
                                            h3_final, opt_dev=opt_dev)
                log_plotting_data(opt_dev, title_prefix="%d %d %d" % (h1_final,
                                                                      h2_final,
                                                                      h3_final))
                sleep(2)


if __name__ == '__main__':
    # run_optimisation_through_ds(True)
    run_looped_optimisation(2, 7, 13)
